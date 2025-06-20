import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO, StringIO
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from sklearn.linear_model import LinearRegression
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill
import smtplib
from email.message import EmailMessage
import datetime
import schedule
import threading
import time
import os
import pickle
import tempfile

# === USER AUTHENTICATION SETUP ===
USER_CREDENTIALS = {
    "admin": {"password": "admin123", "role": "admin"},
    "analyst": {"password": "data456", "role": "analyst"},
    "viewer": {"password": "view789", "role": "viewer"}
}
SESSION_STORAGE_DIR = "user_sessions"
os.makedirs(SESSION_STORAGE_DIR, exist_ok=True)

def authenticate_user():
    st.sidebar.title("🔐 Login")
    username = st.sidebar.text_input("Username")
    password = st.sidebar.text_input("Password", type="password")
    if st.sidebar.button("Login"):
        user = USER_CREDENTIALS.get(username)
        if user and user["password"] == password:
            st.session_state['authenticated'] = True
            st.session_state['username'] = username
            st.session_state['role'] = user['role']
            st.success(f"Welcome, {username} ({user['role']})")
        else:
            st.error("Invalid username or password")

if 'authenticated' not in st.session_state:
    authenticate_user()
    st.stop()

username = st.session_state['username']
user_role = st.session_state['role']

# === APP CONFIG ===
st.set_page_config(page_title="🔬 Advanced Data Analyzer", layout="wide")
st.title(f"📊 Excel & CSV Data Workspace - User: {username} ({user_role})")

EMAIL_ADDRESS = "your_email@example.com"  # configure
EMAIL_PASSWORD = "your_password"          # configure

# === HELPER FUNCTIONS ===

def save_user_session(user, dataframe):
    path = os.path.join(SESSION_STORAGE_DIR, f"{user}_data.pkl")
    with open(path, "wb") as f:
        pickle.dump(dataframe, f)

def load_user_session(user):
    path = os.path.join(SESSION_STORAGE_DIR, f"{user}_data.pkl")
    if os.path.exists(path):
        with open(path, "rb") as f:
            return pickle.load(f)
    return None

def send_email_report(to_email, subject, body, attachment_data=None, attachment_filename="report.xlsx"):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = to_email
    msg.set_content(body)
    if attachment_data:
        attachment_data.seek(0)
        msg.add_attachment(attachment_data.read(),
                           maintype='application',
                           subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           filename=attachment_filename)
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)

def run_schedule():
    while True:
        schedule.run_pending()
        time.sleep(60)

def generate_pdf_report(summary_lines, chart_paths):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(200, 10, "Data Analysis Summary", ln=True, align='C')
    pdf.set_font("Arial", size=10)
    for line in summary_lines:
        pdf.cell(200, 8, line, ln=True)
    for path in chart_paths:
        pdf.add_page()
        pdf.image(path, w=180)
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_buffer = BytesIO(pdf_bytes)
    pdf_buffer.seek(0)
    return pdf_buffer

# === SIDEBAR - DATA INPUT ===
st.sidebar.header("📂 Data Input")
input_method = st.sidebar.radio("Select input method:", ["Upload File", "Paste Data", "Manual Entry", "Load Last Session"])

df = None

if input_method == "Upload File":
    uploaded_file = st.sidebar.file_uploader("Upload Excel or CSV", type=['xlsx', 'csv'])
    if uploaded_file:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file, engine='openpyxl')

elif input_method == "Paste Data":
    pasted = st.sidebar.text_area("Paste CSV-format data here:")
    if pasted:
        try:
            df = pd.read_csv(StringIO(pasted))
        except Exception:
            st.sidebar.error("Invalid CSV data")

elif input_method == "Manual Entry":
    cols_text = st.sidebar.text_input("Column names (comma separated):")
    n_rows = st.sidebar.number_input("Number of rows", 1, 100, 5)
    if cols_text:
        cols = [c.strip() for c in cols_text.split(",")]
        df = pd.DataFrame({col: [""]*n_rows for col in cols})

elif input_method == "Load Last Session":
    df = load_user_session(username)
    if df is None:
        st.sidebar.info("No saved session found for this user.")

if df is not None:

    st.subheader("🧾 Data Editor")
    disabled = user_role == "viewer"
    edited_df = st.data_editor(df, use_container_width=True, num_rows="dynamic", disabled=disabled)

    if user_role != "viewer":
        st.subheader("🔒 Data Validation")
        for col in edited_df.select_dtypes(include='number').columns:
            min_val = float(edited_df[col].min())
            max_val = float(edited_df[col].max())
            val_range = st.slider(f"Valid range for {col}", min_val, max_val, (min_val, max_val))
            edited_df = edited_df[(edited_df[col] >= val_range[0]) & (edited_df[col] <= val_range[1])]
        save_user_session(username, edited_df)

    # === Sidebar - Statistical & Data Operations ===
    st.sidebar.header("📊 Statistical & Data Operations")
    operation = st.sidebar.selectbox("Choose operation:",
                                    ["None",
                                     "Sum", "Mean", "Standard Deviation",
                                     "T-Test", "Chi-Square Test",
                                     "Moving Average",
                                     "Pivot Table",
                                     "Linear Regression"])

    numeric_cols = edited_df.select_dtypes(include='number').columns.tolist()
    categorical_cols = edited_df.select_dtypes(include='object').columns.tolist()

    if operation != "None":
        st.subheader(f"⚙️ {operation} Results")
        if operation in ["Sum", "Mean", "Standard Deviation"]:
            col = st.selectbox(f"Select numeric column for {operation}", numeric_cols)
            if col:
                res = None
                if operation == "Sum":
                    res = edited_df[col].sum()
                elif operation == "Mean":
                    res = edited_df[col].mean()
                elif operation == "Standard Deviation":
                    res = edited_df[col].std()
                st.write(f"{operation} of **{col}**: {res:.4f}")

        elif operation == "T-Test":
            col1 = st.selectbox("Select first numeric column", numeric_cols, key="tt_col1")
            col2 = st.selectbox("Select second numeric column", numeric_cols, key="tt_col2")
            if col1 and col2:
                t_stat, p_val = stats.ttest_ind(edited_df[col1].dropna(), edited_df[col2].dropna())
                st.write(f"T-test between **{col1}** and **{col2}**: t = {t_stat:.4f}, p = {p_val:.4f}")

        elif operation == "Chi-Square Test":
            col1 = st.selectbox("Select first categorical column", categorical_cols, key="chi_col1")
            col2 = st.selectbox("Select second categorical column", categorical_cols, key="chi_col2")
            if col1 and col2:
                table = pd.crosstab(edited_df[col1], edited_df[col2])
                chi2, p, dof, expected = stats.chi2_contingency(table)
                st.write(f"Chi-square between **{col1}** and **{col2}**: chi2 = {chi2:.4f}, p = {p:.4f}")

        elif operation == "Moving Average":
            col = st.selectbox("Select numeric column", numeric_cols)
            window = st.number_input("Window size", 1, len(edited_df), 3)
            if col:
                ma = edited_df[col].rolling(window=window).mean()
                st.line_chart(ma)
                edited_df[f"{col}_MA_{window}"] = ma

        elif operation == "Pivot Table":
            index = st.multiselect("Index (rows)", edited_df.columns.tolist())
            columns = st.multiselect("Columns", edited_df.columns.tolist())
            values = st.multiselect("Values (numeric)", numeric_cols)
            aggfunc = st.selectbox("Aggregation function", ["sum", "mean", "count", "min", "max"])
            if index and values:
                pivot = pd.pivot_table(edited_df, index=index, columns=columns if columns else None,
                                       values=values, aggfunc=aggfunc)
                st.dataframe(pivot)

        elif operation == "Linear Regression":
            y_col = st.selectbox("Dependent variable (Y)", numeric_cols)
            x_cols = st.multiselect("Independent variable(s) (X)", [c for c in numeric_cols if c != y_col])
            if y_col and x_cols:
                X = edited_df[x_cols].dropna()
                y = edited_df.loc[X.index, y_col]
                model = LinearRegression().fit(X, y)
                st.write(f"Coefficients: {dict(zip(x_cols, model.coef_))}")
                st.write(f"Intercept: {model.intercept_:.4f}")
                st.write(f"R² Score: {model.score(X, y):.4f}")

    # === Sidebar - Chart Generator ===
    st.sidebar.header("📈 Chart Generator")
    chart_type = st.sidebar.selectbox("Chart type", ["Bar", "Line", "Pie"])
    x_axis = st.sidebar.selectbox("X-axis", edited_df.columns.tolist())
    y_axis = st.sidebar.selectbox("Y-axis (numeric)", numeric_cols)
    color = st.sidebar.color_picker("Choose chart color", "#0d6efd")

    if st.sidebar.button("Generate Chart"):
        fig, ax = plt.subplots(figsize=(8, 5))
        if chart_type == "Bar":
            data = edited_df.groupby(x_axis)[y_axis].sum()
            data.plot(kind='bar', color=color, ax=ax)
        elif chart_type == "Line":
            data = edited_df.groupby(x_axis)[y_axis].sum()
            data.plot(kind='line', color=color, marker='o', ax=ax)
        elif chart_type == "Pie":
            data = edited_df.groupby(x_axis)[y_axis].sum()
            data.plot(kind='pie', autopct='%1.1f%%', ax=ax, ylabel='', colors=[color]*len(data))
        st.pyplot(fig)

    # === Export Excel with formatting and multiple sheets ===
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        edited_df.to_excel(writer, index=False, sheet_name='Edited Data')
        if 'pivot' in locals():
            pivot.to_excel(writer, sheet_name='Pivot Table')

        wb = writer.book
        ws = wb['Edited Data']
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", fill_type="solid")
    output.seek(0)

    st.download_button("Download Excel with Analysis", data=output,
                       file_name="analysis_output.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # === PDF Report Generation ===
    st.subheader("📄 Generate PDF Summary Report")

    df_summary = [f"Rows: {len(edited_df)}", f"Columns: {len(edited_df.columns)}"]
    for col in numeric_cols:
        df_summary.append(f"{col} - Mean: {edited_df[col].mean():.2f}, Std: {edited_df[col].std():.2f}")

    # Create some temp charts for the PDF
    chart_files = []
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as f1:
        fig, ax = plt.subplots()
        edited_df[numeric_cols].hist(ax=ax)
        plt.tight_layout()
        fig.savefig(f1.name)
        plt.close(fig)
        chart_files.append(f1.name)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as f2:
        fig, ax = plt.subplots(figsize=(6,4))
        sns.heatmap(edited_df[numeric_cols].corr(), annot=True, cmap='coolwarm', ax=ax)
        plt.tight_layout()
        fig.savefig(f2.name)
        plt.close(fig)
        chart_files.append(f2.name)

    pdf_buffer = generate_pdf_report(df_summary, chart_files)
    st.download_button("Download PDF Report", data=pdf_buffer, file_name="summary_report.pdf", mime="application/pdf")

    # === Email sending and scheduling (admin and analyst only) ===
    if user_role in ["admin", "analyst"]:
        st.subheader("📧 Email Report")
        email_address = st.text_input("Recipient Email Address:")
        if st.button("Send Report via Email"):
            try:
                send_email_report(email_address, "Data Analysis Report",
                                  "Attached is your data analysis report.", output)
                st.success("Email sent successfully!")
            except Exception as e:
                st.error(f"Email sending failed: {e}")

        st.subheader("📆 Schedule Daily Email")
        schedule_enabled = st.checkbox("Enable daily email schedule")
        schedule_time = st.time_input("Schedule time", datetime.time(8,0))
        if schedule_enabled and email_address:
            schedule.clear()
            schedule.every().day.at(schedule_time.strftime("%H:%M")).do(
                send_email_report,
                to_email=email_address,
                subject="Scheduled Data Report",
                body="Your scheduled data report is attached.",
                attachment_data=output
            )
            st.info(f"Email scheduled daily at {schedule_time.strftime('%H:%M')} to {email_address}")

            if not hasattr(st.session_state, "scheduler_started"):
                thread = threading.Thread(target=run_schedule, daemon=True)
                thread.start()
                st.session_state.scheduler_started = True

else:
    st.info("Please input or upload data to begin.")
